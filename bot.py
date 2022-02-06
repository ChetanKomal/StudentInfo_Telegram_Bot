import os
import openpyxl as xl
import telebot
import time
my_secret = "YOUR_KEY"
wb = xl.load_workbook("h1copy.xlsx")
ws = wb.active
API_KEY = my_secret
bot = telebot.TeleBot(API_KEY)

def find(msg):
 a = ["0","1","rollno: ","Student Name: ","Father Name: ","Gender: ","Branch: ","Group: ","specialization: ","Mobile no: ","Email: ","Status: "]
 tmp = []
 x=msg
 for row in range(3,368):
     if(int(ws[f"B{row}"].value) == int(x)):
         for i in range(3,12):
             y=chr(64+i)
             tmp.append(str(a[i]) + " " + str(ws[f"{y}{row}"].value))
 return tmp           

def write_to_file(message):
  file = open("log.txt","a+")
  data = ["time: ",time.ctime(),"command: ",message.text,"ID: ",message.from_user.id,"username: ",message.from_user.username,"firstname: ",message.from_user.first_name,"lastname: ",message.from_user.last_name]
  file.write(str(data))
  file.write("\n\n")
  file.close()

@bot.message_handler(commands=["help","start"])
def helper(message):
  write_to_file(message)
  info = ''' 
         This Bot gives information about students of B.E CSE Branch Batch 2019 CUHP. Just Pass a valid rollno.
         \n Bot Commands:\n{valid_university_id}(Eg. 191198xxxx) \n /help \n /start \n /moreinfo\n  \nDeveloped by: @CK
   '''
  bot.send_message(message.chat.id,info,protect_content=True)

@bot.message_handler(commands=["moreinfo"])
def moreinfo(message):
 write_to_file(message)	
 ing = '''
       This bot provides the following information about students:\n
       Student Photo (if available)
       ID: X
       Student name: X
       Father name: X
       Gender: X
       Branch:X
       Group: X
       Specializaton: X
       Mobile No: X
       Email: xx@chitkarauniversity.edu.in
       Status: Dayscholar or Hosteller\n
       
       Developed By: @CK

 '''
 bot.send_message(message.chat.id,ing,protect_content=True)

@bot.message_handler(regexp="191198\d\d\d\d")
def rep(message):
  write_to_file(message)
  if(find(message.text) != []):
   tmp = find(message.text)
   ff = "\n"+"\n" + "ID: " + message.text + "\n" + tmp[0] + "\n"+tmp[1]+"\n"+tmp[2]+"\n"+tmp[3]+"\n"+tmp[4]+"\n"+tmp[5]+"\n"+tmp[6]+"\n"+tmp[7]+"\n"+tmp[8]+"\n" +"\n" + "\n"+"Developed by @CK"
   pic = f"https://hp.chitkara.edu.in//Storage/Images/Student/{str(message.text)}.jpg"
   try:
    bot.send_photo(message.chat.id,photo=pic,caption=ff,protect_content=True)  
   except:
    bot.send_message(message.chat.id,"IMAGE NOT FOUND"+ff,protect_content=True)   
  else:
   bot.send_message(message.chat.id,"No Data Found",protect_content=True)

@bot.message_handler(func=lambda m: True)
def echo_all(message):
	write_to_file(message)
	bot.send_message(message.chat.id,"Incorrect RollNumber",protect_content=True)  

print("Bot running")
try:
    bot.polling() 
except:
    time.sleep(15)
    bot.polling()    

	
▄█▄    █  █▀ 
█▀ ▀▄  █▄█   
█   ▀  █▀▄   
█▄  ▄▀ █  █  
▀███▀    █   
        ▀                
